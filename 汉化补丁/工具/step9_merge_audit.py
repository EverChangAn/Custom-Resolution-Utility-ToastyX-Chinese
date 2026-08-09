# -*- coding: utf-8 -*-
"""step9: 把核查结论合并进 Excel「代码字符串」sheet 的备注列"""
import csv
from openpyxl import load_workbook

AUDIT = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\code_strings_audit.csv'
XLSX = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\CRU翻译对照表.xlsx'

audit = {}
for r in csv.reader(open(AUDIT, encoding='utf-8-sig')):
    audit[r[0]] = r[2]

VERDICT = {
    '确认UI-可翻': '✅ 界面文本',
    '显示格式串-可翻(保留占位符)': '✅ 显示格式串(占位符已保留)',
    '错误弹窗-可翻': '✅ 错误弹窗',
    '数组数据-可翻': '✅ 下拉/列表数据',
    '布局测量-不翻': '🚫 布局测量，不翻',
    '危险-不翻': '🚫 注册表/API/文件用途，不翻',
    '源码未直接出现(可能被拼接)': '⚠️ 源码未直接出现',
    '待核-需人工确认': '⚠️ 见核查报告',
}

wb = load_workbook(XLSX)
ws = wb['代码字符串']
# 备注列 = 第 7 列
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
    txt = row[0].value
    if txt and txt in audit:
        ws.cell(row=row[0].row, column=7).value = VERDICT.get(audit[txt], audit[txt])

wb.save(XLSX)
print('备注列已更新')

# 统计
from collections import Counter
c = Counter(audit.values())
for k, v in c.items():
    print('  %s: %d' % (k, v))
