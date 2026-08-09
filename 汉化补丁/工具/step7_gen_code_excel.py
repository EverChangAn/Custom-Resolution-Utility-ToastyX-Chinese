# -*- coding: utf-8 -*-
"""step7: 代码字符串翻译对照表 → 追加 Excel sheet2"""
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CODE_CSV = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\strings_code.csv'
XLSX = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\CRU翻译对照表.xlsx'

# ---------- 建议译文 ----------
T = {
    'An error occurred while saving. Not all changes were saved.': '保存时出错，并非所有更改都已保存。',
    '%s: %d detailed resolution%s, %d standard resolution%s': '%s：%d 个详细分辨率%s，%d 个标准分辨率%s',
    ' Color formats (no space for deep color) ': ' 色彩格式（无空间存放深色） ',
    'Range limits: %d-%d Hz, %d-%d kHz, %d MHz': '范围限制：%d-%d Hz，%d-%d kHz，%d MHz',
    ' Latency (no space for interlaced info) ': ' 延迟（无空间存放隔行信息） ',
    'Standard resolutions (%d resolution%s)': '标准分辨率（%d 个分辨率%s）',
    'Vertical total calculator (QFT)': '垂直总行数计算器（QFT）',
    ' Content types (no space left) ': ' 内容类型（无剩余空间） ',
    ' 2.1 features (no space left) ': ' 2.1 特性（无剩余空间） ',
    'Tiled display topology (9.x)': '拼接显示拓扑（9.x）',
    ' Display stream compression ': ' 显示流压缩（DSC） ',
    ' TMDS clock (no space left) ': ' TMDS 时钟（无剩余空间） ',
    'VTB-EXT: Video Timing Block': 'VTB-EXT：视频时序块',
    ' Luminance (no space left) ': ' 亮度（无剩余空间） ',
    ' VRR (no space for range) ': ' VRR（无空间存放范围） ',
    ' Features (no space left) ': ' 特性（无剩余空间） ',
    ' Luminance (2 bytes left) ': ' 亮度（剩余 2 字节） ',
    'Custom Resolution Utility': '自定义分辨率工具',
    ' Latency (no space left) ': ' 延迟（无剩余空间） ',
    ' Luminance (1 byte left) ': ' 亮度（剩余 1 字节） ',
    'No detailed resolutions': '无详细分辨率',
    'Actual: %lld.%03lld kHz': '实际：%lld.%03lld kHz',
    'Actual: %lld.%0*lld MHz': '实际：%lld.%0*lld MHz',
    'Data blocks (%d byte%s)': '数据块（%d 字节%s）',
    'No standard resolutions': '无标准分辨率',
    'Default extension block': '默认扩展块',
    '12 slices up to 400 MHz': '12 切片，最高 400 MHz',
    '16 slices up to 400 MHz': '16 切片，最高 400 MHz',
    ' Variable refresh rate ': ' 可变刷新率（VRR） ',
    'Product identification': '产品标识',
    'Tiled display topology': '拼接显示拓扑',
    ' Detailed resolutions ': ' 详细分辨率 ',
    'Actual: %lld.%03lld Hz': '实际：%lld.%03lld Hz',
    ' Standard resolutions ': ' 标准分辨率 ',
    'Failed to import file.': '导入文件失败。',
    '40 Gbps (10 x 4 lanes)': '40 Gbps（10 x 4 通道）',
    '48 Gbps (12 x 4 lanes)': '48 Gbps（12 x 4 通道）',
    '2 slices up to 340 MHz': '2 切片，最高 340 MHz',
    '4 slices up to 340 MHz': '4 切片，最高 340 MHz',
    '8 slices up to 340 MHz': '8 切片，最高 340 MHz',
    '8 slices up to 400 MHz': '8 切片，最高 400 MHz',
    'Vendor-specific video': '厂商自定义视频',
    'Vendor-specific audio': '厂商自定义音频',
    'Color characteristics': '色彩特性',
    'Failed to write file.': '写入文件失败。',
    '18 Gbps (6 x 3 lanes)': '18 Gbps（6 x 3 通道）',
    '24 Gbps (6 x 4 lanes)': '24 Gbps（6 x 4 通道）',
    '32 Gbps (8 x 4 lanes)': '32 Gbps（8 x 4 通道）',
    '1 slice up to 340 MHz': '1 切片，最高 340 MHz',
    ' VRR (no space left) ': ' VRR（无剩余空间） ',
    ' DSC (no space left) ': ' DSC（无剩余空间） ',
    'Vendor-specific data': '厂商自定义数据',
    'HDR dynamic metadata': 'HDR 动态元数据',
    '4:2:0 capability map': '4:2:0 能力映射',
    'Detailed resolutions': '详细分辨率',
    'Dynamic range limits': '动态范围限制',
    '9 Gbps (3 x 3 lanes)': '9 Gbps（3 x 3 通道）',
    'HDR static metadata': 'HDR 静态元数据',
    'Detailed resolution': '详细分辨率',
    'Type 10 resolutions': '类型 10 分辨率',
    'Display device data': '显示设备数据',
    'No extension blocks': '无扩展块',
    'No HDMI resolutions': '无 HDMI 分辨率',
    '%s(%d slot%s left) ': '%s（剩余 %d 个插槽%s） ',
    '%s(%d byte%s left) ': '%s（剩余 %d 字节%s） ',
    'Room configuration': '房间配置',
    'Type 8 resolutions': '类型 8 分辨率',
    'Extension override': '扩展覆盖',
    'Display parameters': '显示参数',
    'Type 2 resolutions': '类型 2 分辨率',
    'Type 3 resolutions': '类型 3 分辨率',
    'Type 4 resolutions': '类型 4 分辨率',
    'Type 5 resolutions': '类型 5 分辨率',
    'Type 6 resolutions': '类型 6 分辨率',
    'Type 9 resolutions': '类型 9 分辨率',
    ' Extension blocks ': ' 扩展块 ',
    ' HDMI resolutions ': ' HDMI 分辨率 ',
    '4:2:0 resolutions': '4:2:0 分辨率',
    'Display interface': '显示接口',
    'Serial number: %s': '序列号：%s',
    'No TV resolutions': '无电视分辨率',
    'No audio formats': '无音频格式',
    'HDMI 2.0 support': 'HDMI 2.0 支持',
    'HDMI 2.1 support': 'HDMI 2.1 支持',
    'Video capability': '视频能力',
    'Video preference': '视频偏好',
    'Speaker location': '扬声器位置',
    'VESA resolutions': 'VESA 分辨率',
    'Power sequencing': '电源时序',
    'CVT-RB2 standard': 'CVT-RB2 标准',
    'Other resolution': '其他分辨率',
    ' TV resolutions ': ' 电视分辨率 ',
    ' Audio formats ': ' 音频格式 ',
    'Sink capability': '接收端能力',
    '%d resolution%s': '%d 个分辨率%s',
    'CTA resolutions': 'CTA 分辨率',
    'CTA data blocks': 'CTA 数据块',
    'CVT-RB standard': 'CVT-RB 标准',
    'Unknown Display': '未知显示器',
    ' Color formats ': ' 色彩格式 ',
    ' Content types ': ' 内容类型 ',
    '30-bit (10 bpc)': '30 位（10 bpc）',
    '36-bit (12 bpc)': '36 位（12 bpc）',
    '42-bit (14 bpc)': '42 位（14 bpc）',
    '48-bit (16 bpc)': '48 位（16 bpc）',
    'Max: %d kbit/s': '最大：%d kbit/s',
    'TV resolutions': '电视分辨率',
    'FreeSync range': 'FreeSync 范围',
    'InfoFrame data': 'InfoFrame 数据',
    'Stereo display': '立体显示',
    '%dx%d (center)': '%dx%d（居中）',
    '%dx%d (bottom)': '%dx%d（底部）',
    '%dx%d (middle)': '%dx%d（中间）',
    'Automatic HDTV': '自动 HDTV',
    'Actual: %s kHz': '实际：%s kHz',
    'Actual: %s MHz': '实际：%s MHz',
    'No data blocks': '无数据块',
    ' 2.1 features ': ' 2.1 特性 ',
    '18-bit (6 bpc)': '18 位（6 bpc）',
    '24-bit (8 bpc)': '24 位（8 bpc）',
    'Audio formats': '音频格式',
    'Speaker setup': '扬声器设置',
    'Serial number': '序列号',
    '%dx%d (right)': '%dx%d（右侧）',
    'Automatic CRT': '自动 CRT',
    'Exact reduced': '精确减少消隐',
    'Actual: %s Hz': '实际：%s Hz',
    ' Data blocks ': ' 数据块 ',
    'Not supported': '不支持',
    'HDMI support': 'HDMI 支持',
    'HDR10+ video': 'HDR10+ 视频',
    '5.1 surround': '5.1 环绕',
    '7.1 surround': '7.1 环绕',
    'Max: %d Mcsc': '最大：%d Mcsc',
    'Max: %d Gbps': '最大：%d Gbps',
    'Range limits': '范围限制',
    'ASCII string': 'ASCII 字符串',
    'Gamma curves': '伽马曲线',
    '%dx%d (left)': '%dx%d（左侧）',
    'Automatic PC': '自动 PC',
    'CVT standard': 'CVT 标准',
    'GTF standard': 'GTF 标准',
    'lines = %d.5': '行数 = %d.5',
    ' TMDS clock ': ' TMDS 时钟 ',
    'Profile: %d': '配置文件：%d',
    'Dolby video': 'Dolby 视频',
    'Colorimetry': '色度学',
    'Dolby audio': 'Dolby 音频',
    'Max: %d MHz': '最大：%d MHz',
    '%dx%d (top)': '%dx%d（顶部）',
    'Native HDTV': '原生 HDTV',
    ' Luminance ': ' 亮度 ',
    ' channel%s': ' 通道%s',
    'HDMI video': 'HDMI 视频',
    'HDMI audio': 'HDMI 音频',
    'Delete all': '全部删除',
    'lines = %d': '行数 = %d',
    '%d Hz (%s)': '%d Hz（%s）',
    'Other (%d)': '其他（%d）',
    ' Features ': ' 特性 ',
    'Other (0)': '其他（0）',
    'Extension': '扩展',
    'Flags: %d': '标志：%d',
    'Level: %d': '级别：%d',
    ' channels': ' 通道',
    'Native PC': '原生 PC',
    'Exact CRT': '精确 CRT',
    ' Latency ': ' 延迟 ',
    'Undefined': '未定义',
    'Extended': '扩展',
    'Name: %s': '名称：%s',
    '%s (%d)': '%s（%d）',
    'Invalid': '无效',
    'Default': '默认',
    ' byte%s': ' 字节%s',
    'Stereo': '立体声',
    'Delete': '删除',
    'Manual': '手动',
    'pixels': '像素',
    'Import': '导入',
    'Export': '导出',
    ' bytes': ' 字节',
    'Other': '其他',
    'Type:': '类型：',
    'Exact': '精确',
    'lines': '行',
    'Error': '错误',
    'Text': '文本',
    'Keep': '保留',
}

# 无需翻译（保留原文）
KEEP = {'MPEG-4 HE AAC + MPS','MPEG-4 AAC LC + MPS','MPEG-4 HE AAC v2','MPEG-4 HE AAC','MPEG-4 AAC LC',
        'MPEG-D USAC','MPEG-H 3D','WMA Pro','LPCM 3D','Auro-Cx','MPEG-1','MPEG-2','AAC LC','E-AC-3',
        'DTS-HD','TrueHD','AC-3','AC-4','DisplayID 1.3','DisplayID 2.0','CTA-861','VTB-EXT',
        'nvapi_QueryInterface','uxtheme.dll','nvapi.dll','CRU_Serial_Number','CRU_Range_Limits',
        'CRU_Extensions','CRU_Name','Default_Monitor','ContainerID','DeviceDesc','MZ','wb','rb','bin',
        'x','OK','Cancel','Hz','MHz','kHz','kB','ms','kbit/s','Mcsc','Gbps','(b)','(1.x)',
        'ADL_Main_Control_Create','ADL_Main_Control_Destroy'}

# 占位/调试/纯格式（跳过，不进表）
SKIP_PATTERNS = ['xcxc', 'hhhh', '9999x9999', 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz',
                 '1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ']

def should_skip(s):
    for p in SKIP_PATTERNS:
        if p in s:
            return True
    if s in ('%dx%d%s%s @ %lld.%03lld Hz (%lld.%0*lld MHz) [%s/%s]%s%s',):
        return False
    # 纯格式串（无字母单词）
    import re
    if not re.search(r'[A-Za-z]{3,}', s):
        return True
    if s.startswith('%') and '%' in s[1:]:
        if not re.search(r'[A-Za-z]{2,}', s):
            return True
    return False

# ---------- 读代码字符串 CSV ----------
rows = list(csv.reader(open(CODE_CSV, encoding='utf-8-sig')))
wb = load_workbook(XLSX)
if '代码字符串' in wb.sheetnames:
    del wb['代码字符串']
ws = wb.create_sheet('代码字符串')
headers = ['序号', '英文原文', '建议译文', 'exe偏移', '编码', '源码文件', '备注']
ws.append(headers)
hfill = PatternFill('solid', fgColor='833C00')
hfont = Font(bold=True, color='FFFFFF')
for c in ws[1]:
    c.fill = hfill
    c.font = hfont
    c.alignment = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

n = 0
filled = 0
skip = 0
for r in rows[1:]:
    txt = r[0]
    offs = r[1]
    enc = r[2]
    files = r[3]
    if should_skip(txt):
        skip += 1
        continue
    zh = T.get(txt, '')
    remark = ''
    if txt in KEEP:
        remark = '保留原文'
    if zh:
        filled += 1
    elif not remark:
        remark = '待定'
    n += 1
    ws.append([n, txt, zh, offs, enc, files, remark])
    row = ws.max_row
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border

widths = [6, 46, 34, 40, 8, 30, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='center', wrap_text=True)
ws.freeze_panes = 'A2'

wb.save(XLSX)
print('已追加 sheet "代码字符串" 到', XLSX)
print('条目:', n, ' 已预填:', filled, ' 跳过(占位/纯格式):', skip)
