# -*- coding: utf-8 -*-
"""step10: 谨慎过滤 — 只保留零风险纯文本翻译，其余全部保留英文
规则：
- dfm sheet：窗体属性（Caption/Text/Hint），纯显示，全部保留预填译文（含 % 的格式串除外）
- 代码 sheet：含 % 格式符的 → 不翻；备注 🚫/⚠️ 的 → 不翻；其余纯文本 → 翻
输出 final_translations.json（回写用）+ 更新 Excel 备注列
"""
import json, re
from openpyxl import load_workbook

XLSX = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\CRU翻译对照表.xlsx'
OUT_JSON = r'C:\Users\Administrator\WorkBuddy\CRU汉化\汉化工作区\final_translations.json'

wb = load_workbook(XLSX)
final = {}          # 原文 -> 译文（最终要写的）
skipped = []        # (原文, 译文, 原因)

# ---------- Sheet1 翻译对照表（dfm 窗体）----------
ws1 = wb['翻译对照表']
n1_keep = n1_skip = 0
for row in ws1.iter_rows(min_row=2):
    en = row[1].value   # 英文原文
    zh = row[2].value   # 建议译文
    if not en:
        continue
    if zh and '%' not in en:
        final[en] = zh
        n1_keep += 1
    else:
        if zh:
            skipped.append((en, zh, '含格式符/无译文'))
            n1_skip += 1
        else:
            n1_skip += 1

# ---------- Sheet2 代码字符串 ----------
ws2 = wb['代码字符串']
n2_keep = n2_skip = 0
for row in ws2.iter_rows(min_row=2):
    en = row[1].value
    zh = row[2].value
    remark = str(row[6].value or '')
    if not en:
        continue
    # 谨慎规则：含 % 格式符 → 不翻
    if zh and '%' not in en and '🚫' not in remark and '⚠️' not in remark and '不翻' not in remark:
        final[en] = zh
        n2_keep += 1
    else:
        if zh:
            reason = []
            if '%' in en:
                reason.append('含格式符')
            if '🚫' in remark or '不翻' in remark:
                reason.append('用途敏感')
            if '⚠️' in remark:
                reason.append('待核')
            skipped.append((en, zh, ';'.join(reason) if reason else '谨慎保留英文'))
        n2_skip += 1

# ---------- 输出 ----------
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(final, f, ensure_ascii=False, indent=1)

print('=== 谨慎过滤结果 ===')
print('Sheet1 dfm: 翻译 %d 条, 保留英文 %d 条' % (n1_keep, n1_skip))
print('Sheet2 代码: 翻译 %d 条, 保留英文 %d 条' % (n2_keep, n2_skip))
print('最终翻译字典: %d 条 → %s' % (len(final), OUT_JSON))
print()
print('=== 代码 sheet 因谨慎保留英文的（含格式符/敏感）===')
for en, zh, reason in skipped[:40]:
    print('  [%s] "%s" → 保留英文' % (reason, en[:60]))
if len(skipped) > 40:
    print('  ... 共 %d 条' % len(skipped))
